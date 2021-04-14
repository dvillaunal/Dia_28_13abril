'''
Datos sacados de:

https://en.wikipedia.org/wiki/Demography_of_the_United_Kingdom

Vamos a exportar un grafico de Lineas en el cual se vera la
población media de Reino Unido
utilizaremos Openpyxl para no generar errores en replit
pero despues trabajremos con pandas y sus dataframes
un de los tipos de diccionarios.

input = .xlsx
output = .png
'''
import openpyxl
import pandas as pd
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt

'''
El paquete ~style~ agrega soporte para "estilos" de trazado fáciles
de cambiar con los mismos parámetros que un archivo rc matplotlib
(que se lee al inicio para configurar Matplotlib).

Hay varios estilos predefinidos proporcionados por Matplotlib.
Por ejemplo, hay un estilo predefinido llamado "ggplot",
que emula la estética de ggplot (un paquete de trazado popular para R ).
Para usar este estilo, simplemente agregue:
'''
mpl.style.use('ggplot')

# Importamos con openpyxl el archivo .xlsx
pm = openpyxl.load_workbook('Población_Media.xlsx')

# dejamos la unica hoja activa y trabajar con ella:
'Nombre de la unica hoja ["Sheet1"]'

h1 = pm['Sheet1']

# Con pandas convertimos a dataframe:
df = pd.DataFrame(h1.values, columns=['Año', 'Población Media'])

# Print al dataframe:
'''
Los comandos:
df.plot()

df.plot(kind = 'line')

crean un gráfico de líneas y los parámetros ~kind = 'line'~ 
le dicen a la función qué datos usar.
Si bien no es necesario pasar
un parámetro ~kind = 'line'~
en el comando para obtener un diagrama de línea,
es mejor agregarlo para un mejor zen de python.  

El primer parámetro, el Año se trazará en el eje x,
y el segundo parámetro, la población media, se trazará en el eje y.
'''
df.plot(x = 'Año', y = 'Población Media', kind='line')

# Damos nombre al grafico:
plt.title('Población Media de Reino Unido')

#Damos nombres a los Ejes:
plt.ylabel('Población Media')
plt.xlabel('Años')

# Exportamos la Imagen en .png:
plt.savefig('PoblacionMedia_UK.png')

#######################################################################

# Gráficos de barras en Pandas con Matplotlib:
'''
Un diagrama de barras es una forma de representar datos
donde la longitud de las barras representa la magnitud / tamaño
de la característica / variable.
Los gráficos de barras suelen representar variables numéricas
y categóricas agrupadas en intervalos.

Los diagramas de barras son más efectivos cuando intenta
visualizar datos categóricos que tienen pocas categorías.
Si tenemos demasiadas categorías,
las barras estarán muy desordenadas en la figura
y serán difíciles de entender. Son buenos para datos categóricos
porque puede ver fácilmente la diferencia entre las categorías según el tamaño de la barra.

Vamos a ver una comparación de algunos de los paises con mejor PIB en el 2018 (en Dolares internacionales):

Fuente:
https://en.wikipedia.org/wiki/List_of_countries_by_GDP_(PPP)_per_capita
'''
# Diagrama de Barras:
'''
Para crear un diagrama de barras usaremos df.plot() nuevamente.
Esta vez podemos pasar uno de dos argumentos a través del parámetro ~kind~ en plot():

-> kind= bar
  crea un gráfico de barras verticales

-> kind=barh
  crea un diagrama de barras horizontales

De manera similar,
el comando df.plot() para el gráfico de barras requerirá tres parámetros:
1. valores de x
2. valores de y
3. tipo de gráfico.
'''
# Leemos el archivo:

di = openpyxl.load_workbook('Dollar_International.xlsx')
hd1 = di['Hoja1']

# Convertimos a data frame:

di = pd.DataFrame(hd1.values, columns=['País', '(PIB)'])

# Grafico de barras horizontal y definimos titulo (Codigo mas Limpio):

'Quitamos unas lineas de codigo con el title dentro del .plot(), además eliminaremos:'
'esa leyenda ya que tenemos solo una grafica no es necesario expresar aquello que esta implicito'
'La eliminaremos con el ~legend = False~'

di.plot(x ='País', y='(PIB)', kind = 'barh',
        color = 'green',
        title = '(PIB) en Dolares Internacionales',
        legend = False)

# Definimos nuestras etiquietas:

plt.ylabel('País')
plt.xlabel('(PIB)')

# Exportamos la imagen:
plt.savefig('PIB_dollarI.png')

################################################################

'''
Por ultimo graficaremos un grafico circular

Un gráfico circular es un gráfico circular que muestra proporciones numéricas
dividiendo un círculo en porciones proporcionales. 

Usaremos un gráfico circular para ilustrar la proporción (porcentaje)
de la población dividida por continentes.
en 2019

datos extraidos de:
https://en.wikipedia.org/wiki/List_of_continents_by_population
'''

# leer el archivo

pc = openpyxl.load_workbook('Poblacion.xlsx')
hc = pc['Hoja1']

# pasar a data frame:
dc = pd.DataFrame(hc.values, columns=['Continente', 'Población'])

# Convertimeos una columna en indice con el metodo .set_index()
dc = dc.set_index('Continente')

# Grafica circular:
'''
Podemos crear gráficos circulares en Matplotlib pasando la palabra clave ~kind=pie~ en df.plot().
'''
# dc.plot(kind = 'pie', y='Población') <- si ejecutamos la linea veremos que quedara sobre puesto las leyendas del pie
'''
por ende utilizaremos ~figsize~

-> figsize(float, float)
  por defecto: rcParams["figure.figsize"](por defecto: )[6.4, 4.8]
  Ancho, alto en pulgadas.
'''
dc.plot(kind = 'pie', y='Población', figsize=(10, 10))
plt.title('Población x Continente')

plt.savefig('PoblacionXcontinente.png')